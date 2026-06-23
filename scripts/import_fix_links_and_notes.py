#!/usr/bin/env python3
"""
Fix 1: Contacts with multiple company IDs (semicolon-separated) — link to first valid company.
Fix 2: Import company Notes as NOTE activities.
"""

import openpyxl
import psycopg2
import uuid
from pathlib import Path

DOWNLOADS = Path.home() / "Downloads"
DB_URL = "postgresql://vijay@localhost:5432/ennablcrm_local"

def gen_id():
    return str(uuid.uuid4())

def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s in ("", "(No value)", "N/A", "None") else s

conn = psycopg2.connect(DB_URL)
conn.autocommit = False
cur = conn.cursor()

# ── Load hubspot_id → CRM company id map ─────────────────────────────────────
cur.execute("SELECT hubspot_id, id FROM companies WHERE hubspot_id IS NOT NULL")
hs_company_to_crmid = {row[0]: row[1] for row in cur.fetchall()}
print(f"Loaded {len(hs_company_to_crmid)} company hubspot→crm mappings")

# ── FIX 1: contacts with multiple company IDs ─────────────────────────────────
print("\n── Fix 1: Multi-company-ID contacts ──")

wb = openpyxl.load_workbook(DOWNLOADS / "All Contacts_18June2026.xlsx", read_only=True, data_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
headers = [str(h).strip() if h else "" for h in rows[0]]

# Load hubspot_id → CRM contact id
cur.execute("SELECT id FROM contacts LIMIT 1")  # just to verify table exists
cur.execute("""
    SELECT c.id, co.hubspot_id
    FROM contacts c
    JOIN companies co ON c.company_id = co.id
    WHERE co.name = '(unnamed)'
    LIMIT 1000
""")
# Build contact email → crm id for updating
cur.execute("SELECT id, email FROM contacts WHERE email IS NOT NULL")
contact_email_to_id = {row[1]: row[0] for row in cur.fetchall()}

fixed = 0
for r in rows[1:]:
    row = dict(zip(headers, r))
    raw_ids = clean(row.get("Associated Company IDs")) or clean(row.get("Associated Company"))
    if not raw_ids:
        continue

    # Handle semicolon-separated IDs
    ids = [x.strip() for x in str(raw_ids).split(";") if x.strip()]
    if len(ids) <= 1:
        continue  # already handled by main import

    # Try each ID until we find one that matches a company
    company_crm_id = None
    for hs_id in ids:
        if hs_id in hs_company_to_crmid:
            company_crm_id = hs_company_to_crmid[hs_id]
            break

    if not company_crm_id:
        continue

    # Find this contact's CRM id by email
    email = clean(row.get("Email"))
    if email and email in contact_email_to_id:
        contact_crm_id = contact_email_to_id[email]
        cur.execute(
            "UPDATE contacts SET company_id = %s WHERE id = %s AND (company_id IS NULL OR company_id IN (SELECT id FROM companies WHERE name = '(unnamed)'))",
            (company_crm_id, contact_crm_id)
        )
        if cur.rowcount > 0:
            fixed += 1

conn.commit()
wb.close()
print(f"  ✓ Fixed {fixed} contacts with multiple company IDs")

# ── FIX 2: Import company Notes as NOTE activities ─────────────────────────────
print("\n── Fix 2: Company notes ──")

# Load vijay's user id to attribute notes to
cur.execute("SELECT id FROM users WHERE email = 'vsrinivasan@ennabl.com' LIMIT 1")
row = cur.fetchone()
vijay_id = row[0] if row else None

wb2 = openpyxl.load_workbook(DOWNLOADS / "All Companies_18June2026.xlsx", read_only=True, data_only=True)
ws2 = wb2.active
rows2 = list(ws2.iter_rows(values_only=True))
headers2 = [str(h).strip() if h else "" for h in rows2[0]]

inserted_notes = 0
for r in rows2[1:]:
    row = dict(zip(headers2, r))
    hs_id = clean(row.get("HubSpot Record ID"))
    note_text = clean(row.get("Notes"))
    if not note_text or not hs_id:
        continue

    company_crm_id = hs_company_to_crmid.get(str(hs_id))
    if not company_crm_id:
        continue

    # Insert as a NOTE activity on the company
    try:
        cur.execute("""
            INSERT INTO activities (id, type, title, description, user_id, company_id, created_at)
            VALUES (%s, 'NOTE', %s, %s, %s, %s, now())
        """, (
            gen_id(),
            note_text[:60] if len(note_text) > 60 else note_text,
            note_text,
            vijay_id,
            company_crm_id,
        ))
        inserted_notes += 1
    except Exception as e:
        conn.rollback()
        continue

    if inserted_notes % 100 == 0:
        conn.commit()

conn.commit()
wb2.close()
print(f"  ✓ Imported {inserted_notes} company notes")

# ── VERIFY ────────────────────────────────────────────────────────────────────
print("\n── Final counts ──")
cur.execute("SELECT COUNT(*) FROM contacts WHERE company_id IS NOT NULL")
print(f"  Contacts with company: {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM contacts c JOIN companies co ON c.company_id=co.id WHERE co.name != '(unnamed)'")
print(f"  Contacts linked to named company: {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM activities WHERE type='NOTE'")
print(f"  Notes (activities): {cur.fetchone()[0]}")

cur.close()
conn.close()
print("\nDone.")
